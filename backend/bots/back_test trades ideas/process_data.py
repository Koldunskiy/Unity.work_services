import pandas as pd
import re
from datetime import datetime, timedelta
import json
import requests
import pytz
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from settings import TOKEN, TG_CH, PARAMETERS

'''Рабочие значения токена'''
TOKEN = TOKEN
TG_CH = TG_CH
URL = 'https://api.telegram.org/bot'


def send_message(text: str, mod: int = 0) -> None:
    """Отправляет или редактирует сообщение в Telegram.

    Args:
        text (str): Текст сообщения.
        mod (int, optional): Режим (0 - новое сообщение, 1 - редактирование). Defaults to 0.
    """
    global last_message_id
    ''' {0:'Print_new_massage', 1:'Edit_massage'} '''

    if mod == 1:
        # Редактируем последнее отправленное сообщение
        res = requests.post('https://api.telegram.org/bot{}/editMessageText'.format(TOKEN),
                            data={'chat_id': TG_CH, 'message_id': last_message_id, 'text': text})
    if mod == 0:
        # last_message_id = None
        # Отправляем новое сообщение
        res = requests.get('https://api.telegram.org/bot{}/sendMessage'.format(TOKEN),
                           params={'chat_id': TG_CH, 'text': text})
        last_message_id = res.json().get('result', {}).get('message_id')


PARAMETERS = PARAMETERS

with open('right_instumets.json', 'r', encoding='utf-8') as f:
    RIGHT_INSTRUMENTS = json.load(f)

'ex = excel = dataframe'


def processing_excel_data(dataframe_name: str) -> pd.DataFrame:
    """Обрабатывает данные из Excel-файла.

    Удаляет пропуски, проверяет завершенность по времени и приводит к численным значениям.

    Args:
        dataframe_name (str): Путь к Excel-файлу.

    Returns:
        pd.DataFrame: Обработанный DataFrame.
    """
    '''Данная функция обрабатывает Excel файл. Удаляет пропуски, проверяет на то, что все иеди по аремени завершилис и приводит к значения к численным данным'''
    ex = pd.read_excel(dataframe_name)
    ex.dropna(inplace=True)
    pattern = r'\s*\(UTC\+\d+\)'
    ex['CREATE TIME'] = ex['CREATE TIME'].apply(lambda row: (re.sub(pattern, '', row)))
    ex['EXPIRATION TIME'] = ex['EXPIRATION TIME'].apply(lambda row: (re.sub(pattern, '', row)))
    # Преобразуем столбец 'EXPIRATION TIME' в объекты datetime
    ex['Help_time'] = pd.to_datetime(ex['EXPIRATION TIME'], format="%d.%m.%Y %H:%M:%S")
    # Добавляем столбец 'mask'
    ex['mask'] = datetime.now() >= ex['Help_time']
    # Фильтруем DataFrame по условию
    ex = ex.loc[ex['mask']]
    # Удаляем столбец 'mask', если он больше не нужен
    ex = ex.drop(columns=['mask', 'Help_time', 'STATUS', 'DESCRIPTION TARGET', 'DESCRIPTION BACKGROUND'])
    ex['INSTRUMENT'] = ex['INSTRUMENT'].apply(lambda x: x.split('.')[0])
    ex[["STOP LOSS", "PRICE", "TAKE PROFIT"]] = ex[["STOP LOSS", "PRICE", "TAKE PROFIT"]].apply(pd.to_numeric,
                                                                                                errors='coerce')
    ex = ex.reset_index(drop=True)
    return ex
    '''Все корректно работает'''


def load_data(symbol_id: str, CREATE_TIME: str, EXPIRATION_TIME: str) -> list:
    """Загружает исторические данные с Exante API.

    Args:
        symbol_id (str): ID инструмента.
        CREATE_TIME (str): Время создания.
        EXPIRATION_TIME (str): Время экспирации.

    Returns:
        list: OHLC-данные.
    """

    def transfrm_to_TIMESTAMP(date_string: str) -> str:
        date = datetime.strptime(date_string, '%d.%m.%Y %H:%M:%S')
        return int(date.timestamp()) * 1000

    start_time = transfrm_to_TIMESTAMP(CREATE_TIME)
    end_time = transfrm_to_TIMESTAMP(EXPIRATION_TIME)
    duration = 60
    size = int(round(((end_time - start_time) / 1000 / duration), 1))
    url = f'https://api-live.exante.eu/md/2.0/ohlc/{symbol_id}/{duration}'
    params = {
        'from': start_time,
        'to': end_time,
        'size': size,
        'type': 'quotes'
    }
    response = requests.get(url, auth=(PARAMETERS["API Key"], PARAMETERS["Secret Key"]), params=params)
    data = json.loads(response.text)
    return data


def main_data_processing(FILE_NAME: str) -> None:
    """Основная обработка данных из файла.

    Тестирует стратегии и сохраняет результаты в Excel.

    Args:
        FILE_NAME (str): Путь к файлу.
    """
    df = processing_excel_data(FILE_NAME)
    strategy_result = []
    send_message(f'Обработанно: 0% \nОставшееся время обработки - {len(df)} мин', mod=0)
    for index, row in df.iterrows():
        INSTRUMENT, SIDE, OPEN_ORDER_TYPE, STOP_LOSS, PRICE, TAKE_PROFIT, CREATE_TIME, EXPIRATION_TIME = row[
            ['INSTRUMENT', 'SIDE',
             'OPEN ORDER TYPE', 'STOP LOSS',
             'PRICE', 'TAKE PROFIT',
             'CREATE TIME',
             'EXPIRATION TIME']]
        if INSTRUMENT in RIGHT_INSTRUMENTS:
            symbol_id = RIGHT_INSTRUMENTS[INSTRUMENT]
            data = load_data(symbol_id, CREATE_TIME, EXPIRATION_TIME)
            result = Strategy_Tester(data, SIDE, OPEN_ORDER_TYPE, STOP_LOSS, PRICE, TAKE_PROFIT)
            strategy_result.append(result)
            time.sleep(60)

        else:
            strategy_result.append(('Данного символа нет в списке инструментов (right_instumets.json)',))
        send_message(
            f'Обработанно: {(index + 1) / len(df) * 100:.1f}% \nОставшееся время обработки - {len(df) - index - 1} мин',
            mod=1)
    df3 = pd.DataFrame(strategy_result)
    new_columns = ['Result', 'CLOSE RATE', 'PL', 'OPEN TIME', 'CLOSE TIME', 'DURATION (H)', 'OPEN TIME IN MLSK',
                   'CLOSE TIME IN MLSK']
    # Переименование столбцов
    df3.columns = new_columns
    f = pd.concat([df, df3], axis=1)
    f.to_excel(FILE_NAME, index=False)


def Strategy_Tester(
        data: list,
        side: str,
        open_order_type: str,
        stop_loss: float,
        price: float,
        take_profit: float
) -> tuple:
    """Тестирует стратегию на данных (аналогично в auxiliary_files).

    Args:
        data (list): OHLC-данные.
        side (str): 'Buy' или 'Sell'.
        open_order_type (str): 'Limit' или 'Stop'.
        stop_loss (float): Стоп-лосс.
        price (float): Цена входа.
        take_profit (float): Тейк-профит.

    Returns:
        tuple: Результат теста.
    """
    df = pd.DataFrame(data)
    # Преобразование timestamp в формат datetime с учетом часового пояса UTC
    df['timestamp2'] = pd.to_datetime(df['timestamp'], unit='ms').dt.tz_localize('UTC')
    # Преобразуем время в часовой пояс UTC+4
    df['timestamp_utc4'] = df['timestamp2'].dt.tz_convert(pytz.timezone('Etc/GMT-3'))
    # Преобразуем дату в формат "d-m-Y H:M:S" и создадим новый столбец 'formatted_date'
    df['formatted_date'] = df['timestamp_utc4'].dt.strftime('%d-%m-%Y %H:%M:%S')
    df = df.drop(columns=["timestamp2", "timestamp_utc4"])
    df[["close", "high", "low", "open"]] = df[["close", "high", "low", "open"]].apply(pd.to_numeric, errors='coerce')
    df = df[::-1]
    df = df.reset_index(drop=True)
    PRICE_INDICATOR = False
    STOP_LOSS_INDICATOR = False
    TAKE_PROFIT_INDICATOR = False
    # 1. Проверяем вход по цене
    if side.capitalize() == 'Buy':
        Flag = True
        if open_order_type.capitalize() == 'Limit':
            price_open = df[df['low'] <= price]
        else:
            price_open = df[df['high'] >= price]
    else:
        Flag = False
        if open_order_type.capitalize() == 'Limit':
            price_open = df[df['high'] >= price]
        else:
            price_open = df[df['low'] <= price]

    if not price_open.empty:
        index_price = price_open.index[0]
        new_df = df.iloc[index_price:].reset_index(drop=True)
        PRICE_INDICATOR = True
        approximate_time_enter, accurate_time_enter = price_open['formatted_date'].iloc[0], \
        price_open['timestamp'].iloc[0]

    # 2. Проверяем стоп-лосс
    if Flag:
        stop_loss_res = new_df[new_df['low'] <= stop_loss]
    else:
        stop_loss_res = new_df[new_df['high'] >= stop_loss]

    if not stop_loss_res.empty:
        index_stop_loss = stop_loss_res.index[0]
        new_df = new_df.iloc[:index_stop_loss + 1].reset_index(drop=True)
        STOP_LOSS_INDICATOR = True
        approximate_time_ST, accurate_time_ST = stop_loss_res['formatted_date'].iloc[0], \
        stop_loss_res['timestamp'].iloc[0]

    # 3. Проверяем на take profit
    if Flag:
        take_profit_res = new_df[new_df['high'] >= take_profit]
    else:
        take_profit_res = new_df[new_df['low'] <= take_profit]

    if not take_profit_res.empty:
        take_profit_index = take_profit_res.index[0]
        TAKE_PROFIT_INDICATOR = True
        approximate_time_TP, accurate_time_TP = take_profit_res['formatted_date'].iloc[0], \
        take_profit_res['timestamp'].iloc[0]

    def timedelta(str1: str, str2: str) -> str:
        """Вычисляет разницу между датами (аналогично выше)."""
        date_format = '%d-%m-%Y %H:%M:%S'
        date1 = datetime.strptime(str1, date_format)
        date2 = datetime.strptime(str2, date_format)
        # Вычисление разницы между датами
        time_difference = date2 - date1
        # Извлечение часов, минут и секунд из разницы
        hours, remainder = divmod(time_difference.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        # Форматирование результатов
        result_string = '{:02}:{:02}:{:02}'.format(hours, minutes, seconds)
        return result_string

    # 4. Вывод
    if PRICE_INDICATOR:
        if TAKE_PROFIT_INDICATOR:
            return ("TP", take_profit, abs(take_profit - price), approximate_time_enter, approximate_time_TP,
                    timedelta(approximate_time_enter, approximate_time_TP), accurate_time_enter, accurate_time_TP)
        elif STOP_LOSS_INDICATOR and not TAKE_PROFIT_INDICATOR:
            return ("SL", stop_loss, -1 * abs(stop_loss - price), approximate_time_enter, approximate_time_ST,
                    timedelta(approximate_time_enter, approximate_time_ST), accurate_time_enter, accurate_time_ST)
        else:
            last_price = new_df['close'].iloc[-1]
            approximate_time_ST = new_df['formatted_date'].iloc[-1]
            accurate_time_ST = new_df['timestamp'].iloc[-1]
            return ("EXP",
                    last_price,
                    last_price - price if side == 'Buy' else price - last_price,
                    approximate_time_enter,
                    approximate_time_ST,
                    timedelta(approximate_time_enter, approximate_time_ST),
                    accurate_time_enter, accurate_time_ST)
    else:
        return ('NOT ENTER',)


def vizual(FILE_NAME: str) -> None:
    """Визуализирует данные в Excel (закрашивает строки по статусу).

    Args:
        FILE_NAME (str): Путь к файлу.
    """
    # Загружаете книгу Excel
    workbook = load_workbook(FILE_NAME)
    # Выбираете нужный лист (в данном случае, первый лист)
    sheet = workbook.active
    # Проходитесь по строкам, начиная с второй строки (предполагается, что первая строка - заголовки)
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Проверяете значение в столбце L (12-й столбец, так как индексация начинается с 1)
        if row[11] == "SL":
            # Если значение равно "SL", закрашиваете строку в красный цвет
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_number, column=col_idx)
                cell.fill = PatternFill(start_color="FF4040", end_color="FF4040", fill_type="solid")
                cell.border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))
        if row[11] == "TP":
            # Если значение равно "TL", закрашиваете строку в green цвет
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_number, column=col_idx)
                cell.fill = PatternFill(start_color="9FEE00", end_color="9FEE00", fill_type="solid")
                cell.border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))
        if row[11] == "EXP":
            # Если значение равно "EXP", закрашиваете строку в blue цвет
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_number, column=col_idx)
                cell.fill = PatternFill(start_color="66A3D2", end_color="66A3D2", fill_type="solid")
                cell.border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))
        if row[11] == "NOT":
            # Если значение равно "NOT", закрашиваете строку в желтый цвет
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_number, column=col_idx)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Устанавливаете границы ячейки
                cell.border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))

    # Сохраняете изменения
    workbook.save(FILE_NAME)
    workbook.close()